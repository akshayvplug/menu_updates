import mysql.connector
from get_xml import Menu
from extract import parse_xml
from typing import Text
from zipfile import ZipFile
from connection import db_connection
import logging
import json
from collections import defaultdict
import csv
from openpyxl import Workbook
import os
import pandas as pd

def fetch_table_columns(cursor, table_name):
    # Function to fetch the column names from the existing table in the database
    cursor.execute(f"DESCRIBE {table_name}")
    columns = [row[0] for row in cursor.fetchall()]
    return columns


def generate_report(rest_abbr, id_list):
    # Define the maximum constraint for the 'name' column
    name_max_length = 100
    
    for id in id_list:
        table_name = f"menu_extracts.{rest_abbr.lower()}_{id}"
        
        # Download the xml file from S3 and keep in menu_files
        menu = Menu(rest_abbr, id)
        
        # Extract the zip file and parse the xml file to extract the data
        with ZipFile(f"menu_files/{menu.file_name}", 'r') as zip_f:
            extracted_data = parse_xml(zip_f.open(f"{menu.vendorid}.xml"), id)
        
        print("Data extracted")
        
        try:
            with db_connection() as cursor:
                
                # Generate report on differences (for illustration)
                # You can customize this based on your specific reporting needs
                report = {
                    'new_items': [],
                    'updated_items': [],
                    'deleted_items': [],
                }
                
                # Fetch existing data from the table
                cursor.execute(f"SELECT * FROM {table_name}")
                existing_data = cursor.fetchall()
                
                # Compare with extracted_data to identify differences
                existing_ids = {item['id'] for item in existing_data}
                new_ids = {item['id'] for item in extracted_data}
                
                for item in extracted_data:
                    if item['id'] not in existing_ids:
                        report['new_items'].append(item)
                    else:
                        # Check if item is updated
                        existing_item = next((e for e in existing_data if e['id'] == item['id']), None)
                        if existing_item and existing_item != item:
                            report['updated_items'].append({'old': existing_item, 'new': item})
                
                for item in existing_data:
                    if item['id'] not in new_ids:
                        report['deleted_items'].append(item)
                
                # print("Report on differences:")
                # print(json.dumps(report, indent=2))


                # Generate concise report with grouped product names
                # Generate concise report with grouped product names
                concise_report = {
                    'new_items': defaultdict(list),
                    'updated_items': defaultdict(list),
                    'deleted_items': defaultdict(list),
                }
                
                for item in report['new_items']:
                    if 'id' in item and 'chain_id' in item and 'is_modifier' in item and 'name' in item and 'is_available' in item:
                        concise_report['new_items'][item['product_name']].append(item)
                
                for item in report['updated_items']:
                    if 'old' in item and 'new' in item and 'id' in item['new'] and 'chain_id' in item['new'] and 'is_modifier' in item['new'] and 'name' in item['new'] and 'is_available' in item['new']:
                        concise_report['updated_items'][item['new']['product_name']].append({'old': item['old'], 'new': item['new']})
                
                for item in report['deleted_items']:
                    if 'id' in item and 'chain_id' in item and 'is_modifier' in item and 'name' in item and 'is_available' in item:
                        concise_report['deleted_items'][item['product_name']].append(item)
                
                # Create Excel workbook
                wb = Workbook()
                ws_new = wb.create_sheet(title='new_items')
                ws_updated = wb.create_sheet(title='updated_items')
                ws_deleted = wb.create_sheet(title='deleted_items')
                
                # Formatting new_items by product_name
                ws_new.append(['Product Name', 'ID', 'Chain ID', 'Is Modifier', 'Name', 'Is Available'])
                new_items_grouped = {}
                for item in report['new_items']:
                    product_name = item['product_name']
                    if product_name not in new_items_grouped:
                        new_items_grouped[product_name] = []
                    new_items_grouped[product_name].append(item)
                
                for product_name, items in new_items_grouped.items():
                    for item in items:
                        ws_new.append([product_name, item['id'], item['chain_id'], item['is_modifier'], item['name'], item['is_available']])
                
                # Formatting updated_items by product_name
                ws_updated.append(['Product Name', 'Old ID', 'Old Chain ID', 'Old Is Modifier', 'Old Name', 'Old Is Available',
                                   'New ID', 'New Chain ID', 'New Is Modifier', 'Name', 'New Is Available'])
                for update in report['updated_items']:
                    old_item = update['old']
                    new_item = update['new']
                    ws_updated.append([new_item['product_name'], old_item['id'], old_item['chain_id'], old_item['is_modifier'], old_item['name'], old_item['is_available'],
                                       new_item['id'], new_item['chain_id'], new_item['is_modifier'], new_item['name'], new_item['is_available']])
                
                # Formatting deleted_items
                ws_deleted.append(['ID', 'Chain ID', 'Is Modifier', 'Name', 'Is Available', 'Product Name'])
                for item in report['deleted_items']:
                    ws_deleted.append([item['id'], item['chain_id'], item['is_modifier'], item['name'], item['is_available'], item['product_name']])
                
                # Save Excel workbook
                excel_filename = f"output/summary_report_{rest_abbr}_{id}.xlsx"
                wb.save(excel_filename)
                print(f"Excel file saved: {excel_filename}")
                
                print("Report on differences:")
                print("New Items:", concise_report['new_items'])
                print("Updated Items:", concise_report['updated_items'])
                print("Deleted Items:", concise_report['deleted_items'])


                # Save report to JSON file
                report_filename = f"output/report_{rest_abbr}_{id}.json"
                with open(report_filename, 'w') as report_file:
                    json.dump(report, report_file, indent=2)
                    print(f"Report saved to {report_filename}")

                # Save report to JSON file
                concise_report_filename = f"output/concise_report_{rest_abbr}_{id}.json"
                with open(concise_report_filename, 'w') as report_file:
                    json.dump(concise_report, report_file, indent=2)
                    print(f"Report saved to {report_filename}")
                sql_queries = []
                
                for item in report['new_items']:
                    sql_queries.append(f"INSERT INTO {table_name} (id, chain_id, is_modifier, name, is_available, product_name) " +
                                       f"VALUES ('{item['id']}', '{item['chain_id']}', {item['is_modifier']}, '{item['name']}', {item['is_available']}, '{item['product_name']}')")
                
                for item in report['updated_items']:
                    sql_queries.append(f"UPDATE {table_name} SET " +
                                       f"id = '{item['new']['id']}', chain_id = '{item['new']['chain_id']}', " +
                                       f"is_modifier = {item['new']['is_modifier']}, name = '{item['new']['name']}', " +
                                       f"is_available = {item['new']['is_available']}, product_name = '{item['new']['product_name']}' " +
                                       f"WHERE id = '{item['old']['id']}'")
                
                for item in report['deleted_items']:
                    sql_queries.append(f"DELETE FROM {table_name} WHERE id = '{item['id']}'")
                
                # Save SQL queries to file
                sql_filename = f"output/sql_queries_{rest_abbr}_{id}.sql"
                with open(sql_filename, 'w') as sql_file:
                    sql_file.write("\n".join(sql_queries))
                    print(f"SQL queries saved to {sql_filename}")

                
        
        except mysql.connector.Error as error:
            print(f"Error: {error}")




import os
import pandas as pd
from collections import defaultdict

def amalgamate_excel_reports(output_folder='output'):
    # Get all Excel files in the output folder
    excel_files = [f for f in os.listdir(output_folder) if f.endswith('.xlsx') and 'combined_summary_report.xlsx' not in f]

    if not excel_files:
        print("No Excel files found in the output folder.")
        return

    # Dictionaries to hold data for new_items, updated_items, deleted_items
    new_items_data = defaultdict(lambda: {'stores': set()})
    updated_items_data = defaultdict(lambda: {'stores': set()})
    deleted_items_data = defaultdict(lambda: {'stores': set()})

    # Process each file
    for excel_file in excel_files:
        file_path = os.path.join(output_folder, excel_file)
        
        with pd.ExcelFile(file_path) as xls:
            store_name = excel_file.split('_')[3].replace('.xlsx', '')  # Extract store name from the file name
            sheets = xls.sheet_names
            sheets_to_process = [sheet for sheet in sheets if sheet not in ['Sheet'] and pd.read_excel(xls, sheet_name=sheet).empty is False]
            for sheet_name in sheets_to_process:
            # Read each sheet
                try:
                    if sheet_name in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        
                        # Add store name to the relevant data structure
                        for index, row in df.iterrows():
                            item_key = (row['Name'])  # Unique key by Product Name and Chain ID
                            if sheet_name == 'new_items':
                                new_items_data[item_key]['stores'].add(store_name)
                            elif sheet_name == 'updated_items':
                                updated_items_data[item_key]['stores'].add(store_name)
                            elif sheet_name == 'deleted_items':
                                deleted_items_data[item_key]['stores'].add(store_name)
                except:
                    pass
    # Convert defaultdict to DataFrame for each category
    def create_dataframe(data_dict, sheet_name):
        data = []
        for (product_name), value in data_dict.items():
            stores_list = list(value['stores'])
            data.append({
                'Product Name': product_name,
                'Stores': stores_list
            })
        return pd.DataFrame(data)
    
    # Create a new Excel file with combined data
    combined_excel_file = os.path.join(output_folder, 'combined_summary_report.xlsx')
    with pd.ExcelWriter(combined_excel_file, engine='openpyxl') as writer:
        create_dataframe(new_items_data, 'new_items').to_excel(writer, sheet_name='new_items', index=False)
        create_dataframe(updated_items_data, 'updated_items').to_excel(writer, sheet_name='updated_items', index=False)
        create_dataframe(deleted_items_data, 'deleted_items').to_excel(writer, sheet_name='deleted_items', index=False)
    
    print(f"Combined Excel file saved: {combined_excel_file}")

    # Optionally, delete the original Excel files
    for excel_file in excel_files:
        os.remove(os.path.join(output_folder, excel_file))
        print(f"Deleted file: {excel_file}")

# Example store lists
OLK_STORES = [6155, 6155, 6157, 6144, 6158, 6133,  6140, 30377, 6166, 34285, 6135, 66312, 6162, 6168, 6169, 6146, 6153]
OAK_STORES = [6152, 6167, 6149]

rest_dict = {
    'OLK': OLK_STORES,
    'OAK': OAK_STORES
}

# Generate reports for each store list
for rest_abbr, id_list in rest_dict.items():
    generate_report(rest_abbr=rest_abbr, id_list=id_list)

# Amalgamate the generated reports into a single Excel file
amalgamate_excel_reports()
